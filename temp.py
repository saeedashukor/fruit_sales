from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from pathlib import Path

directory = 'weekly_shop_data'
path = Path(directory).glob('*')
pathname = directory + '/'
files = [file.name for file in path]
temp_file = files[0]
temp_file = temp_file.strip('.xlsx')
temp = temp_file.split('-')
date, shop_code = '-'.join(temp[:3]), '-'.join(temp[3:])
print(date)
print(shop_code)

# for file in files:
#     wb = load_workbook(filename=pathname + file)
#     ws_fruit = wb['by fruit']
#     ws_oh = wb['overheads']
#
#     for row_cells in ws_fruit.iter_rows(min_row=2):
#         fruit, units_bought, cost, units_sold, price, wastage = [cell.value for cell in row_cells]
#
#     for row_cells in ws_oh.iter_rows(min_row=2):
#         personnel, premises, others = [cell.value for cell in row_cells]

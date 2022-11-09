from django.core.management.base import BaseCommand
from openpyxl import load_workbook
import random
from datetime import datetime
import yaml

from shops.models import City, Shop


class Command(BaseCommand):
    def delete_objects(self):
        Shop.objects.all().delete()
        City.objects.all().delete()

    def import_files(self, excel_file, yaml_file):
        # Create City instance by parsing yaml file
        cities = self.create_cities(yaml_file)

        # Read existing workbook (Excel file) to get Shops
        wb = load_workbook(excel_file)
        ws = wb['Sheet']
        self.create_shops(ws)

    def create_cities(self, yaml_file):
        # Read existing yaml file to get Cities
        cities = {}
        with open(yaml_file) as input_file:
            cities_in_file = yaml.load(input_file, Loader=yaml.Loader)['cities']
            for city in cities_in_file:
                if city not in cities:
                    cities[city] = City.objects.create(name=city)
        return cities

    def create_shops(self, ws):
        for row_cells in ws.iter_rows(min_row=2):
            city, shop_name, shop_code = [cell.value for cell in row_cells]
            random_year = random.randrange(start=1980, stop=datetime.now().year)
            Shop.objects.create(city=City.objects.get(name=city),
                                name=shop_name,
                                code=shop_code,
                                address='',
                                postcode='',
                                year_opened=random_year)

    def handle(self, *args, **options):
        response = input(
            "You're about to delete all existing shops & cities, do you wish to continue?\nPress 'y' to continue.\n")
        if response != 'y':
            return

        self.delete_objects()
        self.import_files(excel_file='shops.xlsx', yaml_file='raw_data.yaml')
        self.stdout.write(self.style.SUCCESS('Successfully imported shops & cities into database.'))

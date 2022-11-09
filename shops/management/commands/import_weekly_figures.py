from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import yaml

from shops.models import Fruit, FruitSales, WeeklySales, ShopOverheads, Shop


class Command(BaseCommand):
    def delete_objects(self):
        for model in ShopOverheads, FruitSales, WeeklySales, Fruit:
            model.objects.all().delete()

    def import_files(self, directory, yaml_file):
        source_files = Path(directory).glob('*')
        pathname = directory + '/'
        files = [file.name for file in source_files]
        count = 0

        fruits = self.create_fruits_from_yaml(yaml_file)

        for excel_file in files:
            self.load_files(pathname, excel_file)
            count += 1
            print(f'Imported {count}/{len(files)}: {excel_file}')

    def load_files(self, pathname, excel_file):
        wb = load_workbook(filename=pathname + excel_file)
        ws_fruit = wb['by fruit']
        ws_oh = wb['overheads']

        date, shop_code = self.parse_spreadsheet(excel_file)
        shop = Shop.objects.get(code=shop_code)

        self.create_weekly_sales(date, shop, excel_file)
        weekly_sales = WeeklySales.objects.get(date=date, shop=shop)

        for row_cells in ws_fruit.iter_rows(min_row=2):
            fruit, units_bought, cost, units_sold, price, wastage = [cell.value for cell in row_cells]
            self.create_fruit_sales(weekly_sales, fruit, units_bought, cost, units_sold, price, wastage)

        for row_cells in ws_oh.iter_rows(min_row=2):
            personnel, premises, others = [cell.value for cell in row_cells]
            self.create_overheads(weekly_sales, 'PERS', personnel)
            self.create_overheads(weekly_sales, 'PREM', premises)
            self.create_overheads(weekly_sales, 'OTHER', others)

    def parse_spreadsheet(self, file):
        file_str = file.strip('.xlsx')
        temp = file_str.split('-')
        date, shop_code = '-'.join(temp[:3]), temp[3:][0]
        date = datetime.strptime(date, '%Y-%m-%d')
        return date, shop_code

    def create_fruits_from_yaml(self, yaml_file):
        fruits = {}
        with open(yaml_file) as input_file:
            fruits_in_file = yaml.load(input_file, Loader=yaml.Loader)['fruit']
            for fruit, unit in fruits_in_file:
                if fruit not in fruits:
                    fruits[fruit] = unit
                    Fruit.objects.create(name=fruit, units=unit)
        return fruits

    def create_weekly_sales(self, date, shop, file):
        WeeklySales.objects.create(date=date, shop=shop, file=file)

    def create_fruit_sales(self, weekly_sales, fruit, units_bought, cost, units_sold, price, wastage):
        fruit = Fruit.objects.get(name=fruit)
        FruitSales.objects.create(weekly_sales=weekly_sales,
                                  fruit=fruit,
                                  units_bought=units_bought,
                                  cost_per_unit=cost,
                                  units_sold=units_sold,
                                  price_per_unit=price,
                                  units_waste=wastage)

    def create_overheads(self, weekly_sales, overhead_type, cost):
        ShopOverheads.objects.create(weekly_sales=weekly_sales,
                                     overhead_type=overhead_type,
                                     cost=cost)

    def handle(self, *args, **options):
        response = input(
            "You're about to delete all existing weekly sales data for all shops, do you wish to continue?\nPress 'y' to continue.\n")
        if response != 'y':
            return

        self.delete_objects()
        self.import_files(directory='weekly_shop_data', yaml_file='raw_data.yaml')
        self.stdout.write(self.style.SUCCESS('Successfully imported weekly sales data for all shops into database.'))
